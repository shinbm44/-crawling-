import openpyxl
import requests
from bs4 import BeautifulSoup
wb = openpyxl.Workbook()


ws = wb.create_sheet('data')


 
ws['A1'] = '종목'
ws['B1'] = '현재가'
ws['C1'] = '평균매입가'
ws['D1'] = '잔고수량'
ws['E1'] = '평가금액'
ws['F1'] = '평가수익'
ws['G1'] = '수익률'


codes = [
    '005930',
    '000660',
    '035720'
]

row = 2
for code in codes:
    header = {'User-agent' : 'Mozila/2.0'}
    url = f'https://finance.naver.com/item/sise.naver?code={code}' 
    reponse =  requests.get(url, headers=header)
    html = reponse.text
    soup = BeautifulSoup(html, 'html.parser')
    price = soup.select_one("#_nowVal").text
    price = price.replace(',', '')
    print(price)
    ws[f'B{row}'] = int(price)
    row = row + 1

wb.save(r'C:\Users\user\Desktop\크롤링\selenium\네이버_주식현재가_크롤링\data.xlsx')